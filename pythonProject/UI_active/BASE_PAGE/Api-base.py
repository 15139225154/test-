import json
import requests
from ddt import ddt, data, unpack,file_data
from openpyxl import load_workbook

def openxls():
    workbook = load_workbook('123.xlsx')
    sheet = workbook['Sheet1']
    rows_sheet = sheet.iter_rows()
    #rows_sheet = load_workbook('./Case_xls/case.xlsx')['Sheet1'].iter_rows()
    list1 = []
    for itm in rows_sheet:
        if itm[0].value == 'method':
            continue;
        list = []
        for col in itm:
            list.append(col.value)
        list1.append(list)
    print(list1)
    return list1

def request1(self,method,url,data=None,json=None,headers=None):
    # 发送请求的方法
    if method == 'post':
        # 判断是否使用json来传参（适用于项目中接口参数有使用json传参的）
        return requests.post(url=url, data=data,json=json, headers=headers)
    elif method == 'get':
        return requests.get(url=url, params=data, headers=headers)
    elif method == 'put':
        return requests.put(url=url, data=data, headers=headers)
    elif method == 'delete':
        return requests.delete(url=url,headers=headers)


def API(api,url,data):

    if api == 'get':
        #把数据类型抓转化成字符串
        da = json.dumps(data)
        #把字符传转化成json
        da1 = json.loads(data)
        # 请求头
        headers = {"Content-Type": "application/json;charset=UTF-8"}
        # 执行请求 params=入参
        res = requests.get(url=url, params=da, headers=headers)
        # 打印结果  status_code==状态代码
        print(res.text, res.status_code)

    elif api == 'post':
        da = json.dumps(data)
        # 请求头
        headres = {'Content-Type":"application/json;charset=UTF-8'}
        # 执行请求 params=入参
        res = requests.post(url=url, data=da,json=None, headres=headres)
        print(res.text, res.status_code)


re=openxls()
for i in re:
    print(i[0])
